from openpyxl import load_workbook
import pandas as pd
from pymongo import MongoClient, errors
import mysql.connector
from mysql.connector import Error as MySQLError
import logging
from urllib.parse import unquote, parse_qs, urlparse
import json
from datetime import datetime
import threading
from queue import Queue


def read_excel_to_dict(file_path, columns=None):
    """
    Read columns from Sheet1 of an Excel file and return data in dictionary format.
    
    Args:
        file_path: Path to the Excel file
        columns: List of column names to read. If None, reads all columns
    
    Returns:
        Dictionary where keys are column headers and values are lists of cell values
    
    Example:
        # Read all columns
        data = read_excel_to_dict('input.xlsx')
        
        # Read specific columns
        data = read_excel_to_dict('input.xlsx', columns=['Name', 'Age', 'Email'])
    """
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    worksheet = workbook['Sheet1']
    
    # Read headers from first row
    headers = []
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=1, column=col).value
        if cell_value is not None:
            headers.append(str(cell_value))
        else:
            headers.append(f"Column_{col}")
    
    # Determine which columns to read
    if columns is None:
        # Read all columns
        column_indices = list(range(1, len(headers) + 1))
        column_headers = headers
    else:
        column_indices = []
        column_headers = []
        for col in columns:
            if col in headers:
                col_index = headers.index(col) + 1
                column_indices.append(col_index)
                column_headers.append(col)
    
    # Initialize result dictionary
    result_dict = {}
    for header in column_headers:
        result_dict[header] = []
    
    # Read data from row 2 onwards
    for row in range(2, worksheet.max_row + 1):
        for i, col_index in enumerate(column_indices):
            cell_value = worksheet.cell(row=row, column=col_index).value
            if cell_value is None:
                cell_value = ""
            result_dict[column_headers[i]].append(cell_value)
    
    workbook.close()
    return result_dict


def write_data_to_excel(data_list, output_file='output.xlsx'):
    """
    Write list of dictionaries to Excel file.
    
    Args:
        data_list: List of dictionaries containing the scraped data
        output_file: Path to the output Excel file
    
    Returns:
        None
    """
    if not data_list:
        print("No data to write to Excel")
        return
    
    # Convert list of dictionaries to DataFrame
    df = pd.DataFrame(data_list)
    
    # Write to Excel file
    df.to_excel(output_file, index=False, engine='openpyxl')
    print(f"Data successfully written to {output_file}")
    print(f"Total records: {len(data_list)}")


def load_existing_data(output_file='output.xlsx'):
    """
    Load existing data from Excel file if it exists.
    
    Args:
        output_file: Path to the output Excel file
    
    Returns:
        List of dictionaries containing existing data, empty list if file doesn't exist
    """
    try:
        df = pd.read_excel(output_file, engine='openpyxl')
        return df.to_dict('records')
    except FileNotFoundError:
        print(f"No existing file found at {output_file}. Starting fresh.")
        return []
    except Exception as e:
        print(f"Error loading existing data: {str(e)}. Starting fresh.")
        return []


def get_processed_queries(existing_data):
    """
    Get list of unique queries that have already been processed.
    
    Args:
        existing_data: List of dictionaries containing existing data
    
    Returns:
        Set of processed query strings
    """
    processed_queries = set()
    for item in existing_data:
        search_url = item.get('search_url', '')
        if search_url:
            # Extract query from search_url
            try:
                parsed_url = urlparse(search_url)
                params = parse_qs(parsed_url.query)
                if 'query' in params:
                    query = unquote(params['query'][0])
                    processed_queries.add(query)
            except:
                pass
    return processed_queries


# MongoDB utility functions
class MongoDBHandler:
    def __init__(self, connection_string="mongodb://localhost:27017/", database_name="Jordan", collection_name="sales"):
        """
        Initialize MongoDB connection
        
        Args:
            connection_string: MongoDB connection string
            database_name: Name of the database
            collection_name: Name of the collection
        """
        self.connection_string = connection_string
        self.database_name = database_name
        self.collection_name = collection_name
        self.client = None
        self.db = None
        self.collection = None
        
    def connect(self):
        """Establish connection to MongoDB"""
        try:
            self.client = MongoClient(self.connection_string)
            # Test the connection
            self.client.admin.command('ping')
            self.db = self.client[self.database_name]
            self.collection = self.db[self.collection_name]
            
            # Create unique index on itemId to ensure uniqueness
            try:
                self.collection.create_index("itemId", unique=True)
            except errors.OperationFailure:
                # Index might already exist
                pass
                
            return True
        except Exception as e:
            print(f"❌ Failed to connect to MongoDB: {str(e)}")
            return False
    
    def insert_data_batch(self, data_list):
        """
        Insert batch of data with itemId uniqueness handling
        
        Args:
            data_list: List of dictionaries to insert
            
        Returns:
            Dictionary with insertion results
        """
        if not data_list:
            return {"inserted": 0, "duplicates": 0, "errors": 0}
        
        inserted_count = 0
        duplicate_count = 0
        error_count = 0
        
        for item in data_list:
            try:
                if 'itemId' in item:
                    # Try to insert, will fail if itemId already exists
                    self.collection.insert_one(item)
                    inserted_count += 1
                else:
                    error_count += 1
            except errors.DuplicateKeyError:
                duplicate_count += 1
            except Exception as e:
                error_count += 1
                print(f"❌ Error inserting item {item.get('itemId', 'Unknown')}: {str(e)}")
        
        result = {
            "inserted": inserted_count,
            "duplicates": duplicate_count,
            "errors": error_count
        }
        
        return result
    
    def get_processed_queries_from_db(self):
        """
        Get list of unique queries that have already been processed from MongoDB
        
        Returns:
            Set of processed query strings
        """
        processed_queries = set()
        try:
            # Get distinct search_url values
            search_urls = self.collection.distinct("search_url")
            
            for search_url in search_urls:
                if search_url:
                    try:
                        parsed_url = urlparse(search_url)
                        params = parse_qs(parsed_url.query)
                        if 'query' in params:
                            query = unquote(params['query'][0])
                            processed_queries.add(query)
                    except:
                        pass
        except Exception as e:
            print(f"❌ Error getting processed queries from DB: {str(e)}")
        
        return processed_queries
    
    def get_total_records(self):
        """Get total number of records in the collection"""
        try:
            return self.collection.count_documents({})
        except Exception as e:
            print(f"❌ Error getting record count: {str(e)}")
            return 0
    
    def close_connection(self):
        """Close MongoDB connection"""
        if self.client:
            self.client.close()


def save_data_to_mongodb(data_list, mongo_handler):
    """
    Save list of dictionaries to MongoDB with itemId uniqueness handling
    
    Args:
        data_list: List of dictionaries containing the scraped data
        mongo_handler: MongoDBHandler instance
    
    Returns:
        Dictionary with insertion results
    """
    if not data_list:
        return {"inserted": 0, "duplicates": 0, "errors": 0}
    
    result = mongo_handler.insert_data_batch(data_list)
    return result


# SQL Database utility functions
class SQLDBHandler:
    def __init__(self, host="localhost", port=3306, database="Jordan", user="root", password=None, table_name="sales", pool_size=10):
        """
        Initialize SQL database connection with connection pooling
        
        Args:
            host: Database host
            port: Database port
            database: Database name
            user: Database user
            password: Database password (if None, will prompt for input)
            table_name: Table name
            pool_size: Number of connections in the pool
        """
        self.host = host
        self.port = port
        self.database = database
        self.user = user
        self.password = password
        self.table_name = table_name
        self.pool_size = pool_size
        self.connection_pool = Queue(maxsize=pool_size)
        self.pool_lock = threading.Lock()
        self.connection = None
        self.cursor = None
        
    def _create_connection(self):
        """Create a new database connection"""
        try:
            connection = mysql.connector.connect(
                host=self.host,
                port=self.port,
                database=self.database,
                user=self.user,
                password=self.password,
                charset='utf8mb4',
                collation='utf8mb4_unicode_ci',
                autocommit=False,
                use_unicode=True
            )
            return connection
        except MySQLError as e:
            print(f"❌ Failed to create connection: {str(e)}")
            return None
    
    def _initialize_pool(self):
        """Initialize the connection pool"""
        for _ in range(self.pool_size):
            conn = self._create_connection()
            if conn:
                self.connection_pool.put(conn)
    
    def _get_connection(self):
        """Get a connection from the pool"""
        try:
            # Try to get a connection from pool (non-blocking)
            connection = self.connection_pool.get_nowait()
            
            # Check if connection is still valid
            if connection and connection.is_connected():
                return connection
            else:
                # Connection is dead, create a new one
                return self._create_connection()
        except:
            # Pool is empty, create a new connection
            return self._create_connection()
    
    def _return_connection(self, connection):
        """Return a connection to the pool"""
        if connection and connection.is_connected():
            try:
                self.connection_pool.put_nowait(connection)
            except:
                # Pool is full, close the connection
                connection.close()
        
    def connect(self):
        """Establish connection to SQL database and initialize connection pool"""
        try:
            # If no password provided, prompt for it
            if self.password is None:
                import getpass
                self.password = getpass.getpass(f"Enter MySQL password for user '{self.user}': ")
            
            # First, try to connect to the specific database
            try:
                test_connection = mysql.connector.connect(
                    host=self.host,
                    port=self.port,
                    database=self.database,
                    user=self.user,
                    password=self.password,
                    charset='utf8mb4',
                    collation='utf8mb4_unicode_ci'
                )
                test_connection.close()
            except MySQLError as e:
                if "Unknown database" in str(e):
                    print(f"📋 Database '{self.database}' doesn't exist. Creating it...")
                    
                    # Connect without specifying database to create it
                    temp_connection = mysql.connector.connect(
                        host=self.host,
                        port=self.port,
                        user=self.user,
                        password=self.password,
                        charset='utf8mb4',
                        collation='utf8mb4_unicode_ci'
                    )
                    
                    temp_cursor = temp_connection.cursor()
                    temp_cursor.execute(f"CREATE DATABASE {self.database}")
                    temp_cursor.close()
                    temp_connection.close()
                    
                    print(f"✅ Database '{self.database}' created successfully!")
                else:
                    raise e
            
            # Initialize connection pool
            self._initialize_pool()
            
            # Get one connection for table creation and testing
            self.connection = self._get_connection()
            if self.connection and self.connection.is_connected():
                self.cursor = self.connection.cursor(buffered=True)
                
                # Create table if it doesn't exist
                self.create_table()
                
                print(f"✅ Successfully connected to MySQL database '{self.database}' as '{self.user}'")
                print(f"🔗 Connection pool initialized with {self.pool_size} connections")
                return True
        except MySQLError as e:
            print(f"❌ Failed to connect to MySQL: {str(e)}")
            print(f"💡 Connection details: {self.user}@{self.host}:{self.port}/{self.database}")
            
            # Common troubleshooting tips
            if "Access denied" in str(e):
                print("🔧 Troubleshooting tips:")
                print("   1. Check if your MySQL password is correct")
                print("   2. Make sure MySQL server is running")
                print("   3. Verify the username has access to the database")
                print("   4. Try connecting with MySQL Workbench first to test credentials")
            elif "Unknown database" in str(e):
                print("🔧 Troubleshooting tips:")
                print("   1. Create the database first: CREATE DATABASE Jordan;")
                print("   2. Make sure you're connecting to the right MySQL instance")
            
            return False
    
    def create_table(self):
        """Create the sales table if it doesn't exist"""
        try:
            create_table_query = f"""
            CREATE TABLE IF NOT EXISTS {self.table_name} (
                serialNumber INT AUTO_INCREMENT PRIMARY KEY,
                bids INT,
                cardDescription TEXT,
                cardId VARCHAR(255),
                condition_value VARCHAR(100),
                condition_numeric DECIMAL(3,1),
                date DATETIME,
                date_normal DATE,
                feedback TEXT,
                fixImage TEXT,
                fixImage2 TEXT,
                gemRateId VARCHAR(255),
                gradingCompany VARCHAR(100),
                hitIndex INT,
                hitScore DECIMAL(10,3),
                image TEXT,
                isEbayListing BOOLEAN DEFAULT FALSE,
                itemId VARCHAR(255) UNIQUE NOT NULL,
                listingType VARCHAR(100),
                listPrice DECIMAL(10,2),
                notes TEXT,
                originalImage TEXT,
                platform VARCHAR(100),
                price DECIMAL(10,2),
                psaSpecId VARCHAR(255),
                scrapeSuccess BOOLEAN DEFAULT TRUE,
                search_query TEXT,
                search_url TEXT,
                seller VARCHAR(255),
                setUrl TEXT,
                slabSerial VARCHAR(255),
                thumbnail TEXT,
                title TEXT,
                universalGemRateId VARCHAR(255),
                url TEXT,
                verified BOOLEAN DEFAULT FALSE,
                tier VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_itemId (itemId),
                INDEX idx_search_query (search_query(100)),
                INDEX idx_cardId (cardId),
                INDEX idx_date_normal (date_normal),
                INDEX idx_verified (verified),
                INDEX idx_tier (tier),
                INDEX idx_condition_numeric (condition_numeric),
                INDEX idx_platform (platform),
                INDEX idx_gradingCompany (gradingCompany),
                INDEX idx_created_at (created_at)
            )
            """
            self.cursor.execute(create_table_query)
            self.connection.commit()
        except MySQLError as e:
            print(f"❌ Error creating table: {str(e)}")
    
    def insert_data_batch(self, data_list):
        """
        Insert batch of data with itemId uniqueness handling using individual connection
        
        Args:
            data_list: List of dictionaries to insert
            
        Returns:
            Dictionary with insertion results
        """
        if not data_list:
            return {"inserted": 0, "duplicates": 0, "errors": 0}
        
        # Get a dedicated connection for this batch operation
        connection = self._get_connection()
        if not connection:
            print("❌ Failed to get database connection for batch insert")
            return {"inserted": 0, "duplicates": 0, "errors": len(data_list)}
        
        inserted_count = 0
        duplicate_count = 0
        error_count = 0
        
        try:
            cursor = connection.cursor(buffered=True)
            
            for item in data_list:
                try:
                    if 'itemId' in item and item['itemId']:
                        # Prepare data for insertion
                        insert_data = self._prepare_item_for_sql(item)
                        
                        # Build INSERT IGNORE query to handle duplicates
                        columns = ', '.join(insert_data.keys())
                        placeholders = ', '.join(['%s'] * len(insert_data))
                        
                        insert_query = f"""
                        INSERT IGNORE INTO {self.table_name} ({columns}) 
                        VALUES ({placeholders})
                        """
                        
                        # Execute the query
                        cursor.execute(insert_query, list(insert_data.values()))
                        
                        # Check if row was actually inserted
                        if cursor.rowcount > 0:
                            inserted_count += 1
                        else:
                            duplicate_count += 1
                            
                    else:
                        error_count += 1
                        
                except MySQLError as e:
                    error_count += 1
                    print(f"❌ Error inserting item {item.get('itemId', 'Unknown')}: {str(e)}")
            
            # Commit all changes
            connection.commit()
            cursor.close()
            
        except MySQLError as e:
            print(f"❌ Error in batch insert operation: {str(e)}")
            try:
                connection.rollback()
            except:
                pass
            error_count = len(data_list)
            
        finally:
            # Return connection to pool
            self._return_connection(connection)
        
        result = {
            "inserted": inserted_count,
            "duplicates": duplicate_count,
            "errors": error_count
        }
        
        return result
    
    def _prepare_item_for_sql(self, item):
        """
        Prepare an item dictionary for SQL insertion
        
        Args:
            item: Dictionary containing item data
            
        Returns:
            Dictionary with SQL-compatible data
        """
        # Start with basic required fields
        sql_data = {
            'bids': self._safe_int(item.get('bids')),
            'cardDescription': item.get('cardDescription', ''),
            'cardId': item.get('cardId', ''),
            'condition_value': item.get('condition', ''),
            'condition_numeric': self._parse_condition_numeric(item.get('condition', '')),
            'date': self._safe_datetime(item.get('date')),
            'date_normal': self._safe_date(item.get('date_normal')),
            'feedback': item.get('feedback', ''),
            'fixImage': item.get('fixImage', ''),
            'fixImage2': item.get('fixImage2', ''),
            'gemRateId': item.get('gemRateId', ''),
            'gradingCompany': item.get('gradingCompany', ''),
            'hitIndex': self._safe_int(item.get('hitIndex')),
            'hitScore': self._safe_decimal(item.get('hitScore')),
            'image': item.get('image', ''),
            'isEbayListing': bool(item.get('isEbayListing', False)),
            'itemId': item.get('itemId', ''),
            'listingType': item.get('listingType', ''),
            'listPrice': self._safe_decimal(item.get('listPrice')),
            'notes': item.get('notes', ''),
            'originalImage': item.get('originalImage', ''),
            'platform': item.get('platform', ''),
            'price': self._safe_decimal(item.get('price')),
            'psaSpecId': item.get('psaSpecId', ''),
            'scrapeSuccess': bool(item.get('scrapeSuccess', True)),
            'search_query': self._process_search_query(item.get('search_query', '')),
            'search_url': item.get('search_url', ''),
            'seller': item.get('seller', ''),
            'setUrl': item.get('setUrl', ''),
            'slabSerial': item.get('slabSerial', ''),
            'thumbnail': item.get('thumbnail', ''),
            'title': item.get('title', ''),
            'universalGemRateId': item.get('universalGemRateId', ''),
            'url': item.get('url', ''),
            'verified': bool(item.get('Verified', False)),
            'tier': item.get('Tier', ''),  # Fixed: use 'Tier' (capital T) to match main.py
        }
        
        return sql_data
    
    def _safe_decimal(self, value):
        """Safely convert value to decimal"""
        if value is None or value == '':
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    
    def _safe_int(self, value):
        """Safely convert value to integer"""
        if value is None or value == '':
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None
    
    def _safe_datetime(self, value):
        """Safely convert value to datetime"""
        if value is None or value == '':
            return None
        try:
            # If it's already a datetime object, return as is
            if isinstance(value, datetime):
                return value
            # If it's a string, try to parse it
            if isinstance(value, str):
                # Handle ISO format with T separator
                if 'T' in value:
                    return datetime.fromisoformat(value.replace('Z', '+00:00'))
                # Handle other common formats
                from dateutil import parser
                return parser.parse(value)
            return None
        except (ValueError, TypeError, ImportError):
            return None
    
    def _safe_date(self, value):
        """Safely convert value to date"""
        if value is None or value == '':
            return None
        try:
            # If it's already a date object, return as is
            if hasattr(value, 'date'):
                return value.date() if hasattr(value, 'date') else value
            # If it's a string, try to parse it
            if isinstance(value, str):
                if 'T' in value:
                    # Extract date part from datetime string
                    return datetime.fromisoformat(value.replace('Z', '+00:00').split('T')[0]).date()
                else:
                    return datetime.strptime(value, '%Y-%m-%d').date()
            return None
        except (ValueError, TypeError):
            return None
    
    def _parse_condition_numeric(self, condition_value):
        """
        Parse condition values like 'g8_5' (8.5) or 'g7' (7.0) into numeric format
        
        Args:
            condition_value: String condition value (e.g., 'g8_5', 'g7', 'g10')
            
        Returns:
            float: Numeric condition value or None if cannot parse
        """
        if not condition_value or not isinstance(condition_value, str):
            return None
        
        # Convert to lowercase and strip whitespace
        condition_clean = condition_value.lower().strip()
        
        try:
            # Check if it starts with 'g' (grade)
            if condition_clean.startswith('g'):
                # Remove the 'g' prefix
                numeric_part = condition_clean[1:]
                
                # Handle underscore as decimal separator (g8_5 = 8.5)
                if '_' in numeric_part:
                    parts = numeric_part.split('_')
                    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                        integer_part = int(parts[0])
                        decimal_part = int(parts[1])
                        # Convert g8_5 to 8.5
                        return float(f"{integer_part}.{decimal_part}")
                
                # Handle whole numbers (g7 = 7.0)
                elif numeric_part.isdigit():
                    return float(numeric_part)
                
                # Handle already decimal format (g8.5 = 8.5)
                elif '.' in numeric_part:
                    try:
                        return float(numeric_part)
                    except ValueError:
                        pass
            
            # Try to parse as direct number if no 'g' prefix
            if condition_clean.replace('.', '').replace('_', '').isdigit():
                if '_' in condition_clean:
                    parts = condition_clean.split('_')
                    if len(parts) == 2:
                        return float(f"{parts[0]}.{parts[1]}")
                else:
                    return float(condition_clean)
            
        except (ValueError, IndexError):
            pass
        
        return None
    
    def _process_search_query(self, search_query):
        """
        Process search query by joining words with underscores
        
        Args:
            search_query: Original search query string
            
        Returns:
            str: Search query with words joined by underscores
        """
        if not search_query:
            return ''
        
        # Split by spaces and join with underscores
        # Also handle multiple spaces and strip whitespace
        words = search_query.strip().split()
        return '_'.join(words)
    
    def get_total_records(self):
        """Get total number of records in the table using individual connection"""
        connection = self._get_connection()
        if not connection:
            print("❌ Failed to get database connection for record count")
            return 0
            
        try:
            cursor = connection.cursor()
            count_query = f"SELECT COUNT(*) FROM {self.table_name}"
            cursor.execute(count_query)
            result = cursor.fetchone()
            count = result[0] if result else 0
            cursor.close()
            return count
        except MySQLError as e:
            print(f"❌ Error getting record count: {str(e)}")
            return 0
        finally:
            self._return_connection(connection)
    
    def get_count_for_query(self, query):
        """
        Get the number of records for a specific search query using connection pool
        
        Args:
            query: Search query string
            
        Returns:
            int: Number of records for this query
        """
        connection = self._get_connection()
        if not connection:
            print(f"❌ Failed to get database connection for query count: '{query}'")
            return -1
            
        try:
            cursor = connection.cursor()
            
            # Process the query the same way we do when saving
            processed_query = self._process_search_query(query)
            count_query = f"""
            SELECT COUNT(*) FROM {self.table_name} 
            WHERE search_query = %s
            """
            cursor.execute(count_query, (processed_query,))
            result = cursor.fetchone()
            count = result[0] if result else 0
            cursor.close()
            return count
        except MySQLError as e:
            print(f"❌ Error counting records for query '{query}': {str(e)}")
            return -1
        finally:
            self._return_connection(connection)
    
    def check_item_exists(self, item_id):
        """
        Check if a specific itemId already exists using connection pool
        
        Args:
            item_id: The itemId to check
            
        Returns:
            bool: True if item exists, False otherwise
        """
        connection = self._get_connection()
        if not connection:
            print(f"❌ Failed to get database connection to check item: '{item_id}'")
            return False
            
        try:
            cursor = connection.cursor()
            check_query = f"SELECT COUNT(*) FROM {self.table_name} WHERE itemId = %s"
            cursor.execute(check_query, (item_id,))
            result = cursor.fetchone()
            exists = result[0] > 0 if result else False
            cursor.close()
            return exists
        except MySQLError as e:
            print(f"❌ Error checking if item exists: {str(e)}")
            return False
        finally:
            self._return_connection(connection)
    
    def close_connection(self):
        """Close SQL database connections and clean up pool"""
        try:
            # Close main connection
            if self.cursor:
                self.cursor.close()
            if self.connection and self.connection.is_connected():
                self.connection.close()
            
            # Close all connections in the pool
            while not self.connection_pool.empty():
                try:
                    conn = self.connection_pool.get_nowait()
                    if conn and conn.is_connected():
                        conn.close()
                except:
                    break
                    
            print("✅ All database connections closed successfully")
        except MySQLError as e:
            print(f"❌ Error closing connections: {str(e)}")


def save_data_to_sql(data_list, sql_handler):
    """
    Save list of dictionaries to SQL database with itemId uniqueness handling
    
    Args:
        data_list: List of dictionaries containing the scraped data
        sql_handler: SQLDBHandler instance
    
    Returns:
        Dictionary with insertion results
    """
    if not data_list:
        return {"inserted": 0, "duplicates": 0, "errors": 0}
    
    result = sql_handler.insert_data_batch(data_list)
    return result